#!/usr/bin/env python3
"""One-time generator for the reading-list tables in ``wiki-f26/README.md``.

Reads ``docs/CS858-F26-papers-stripped.xlsx`` (sheet ``UpdatedList``) and prints
the Part 1 / Part 2 sections as HTML tables.

The spreadsheet groups papers under a shared theme. Each theme is rendered as a
full-width section-header row (an HTML ``<th colspan>``) that introduces its
papers; the full-width "Part" banners become Markdown headings above each table.

Per-paper layout, under a single ``Topic`` column and a single ``Reading``
column:

* the assigned reading carries an "Assigned reading" label and links to its
  reading companion when one exists (see ``READY``), otherwise to the shared
  ``under-construction.md`` placeholder with a dagger marker;
* essential readings collapse into a ``<details>`` disclosure beneath the
  primary, keeping the hyperlinks already stored in the spreadsheet (arXiv,
  USENIX, IEEE, ACM, and similar).

Internal links stay relative ``.md`` paths so they resolve when browsing the
wiki repo directly (GitHub, VS Code, Obsidian); the website build rewrites them
to absolute site URLs.

Paper numbers are read directly from the spreadsheet; the sheet is the source
of truth for numbering.

Run from the repo root::

    uv run python3 scripts/build-paper-table.py > /tmp/reading-list.md

The output is spliced into ``wiki-f26/README.md`` by hand so the surrounding prose
stays under editorial control.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import cast

from openpyxl import load_workbook

XLSX = Path("docs/CS858-F26-papers-stripped.xlsx")
SHEET = "UpdatedList"

# Papers whose reading companion already exists, keyed by reading-list number.
READY: dict[int, str] = {
    1: "madry-2018-pgd",
    2: "wei-2023-jailbroken",
    3: "greshake-2023-indirect-prompt-injection",
    4: "qi-2024-shallow-safety-alignment",
    5: "carlini-2022-lira",
    6: "carlini-2021-extracting-training-data",
    7: "abadi-2016-dp-sgd",
    8: "jang-2022-knowledge-unlearning",
    9: "orekondy-2019-knockoff-nets",
    10: "szyller-2019-dawn",
    11: "wang-2019-neural-cleanse",
    12: "zou-2024-poisonedrag",
    13: "zou-2023-representation-engineering",
    14: "duddu-2024-unintended-interactions",
    15: "kirchenbauer-2023-llm-watermark",
    16: "pearce-2023-vulnerability-repair",
    17: "zhang-2025-nexus",
    18: "elatali-2024-blime",
    19: "bao-2025-dp-zo",
    20: "zhang-2024-tee-shielded",
    21: "tang-2024-modelguard",
    22: "moon-2025-asgard",
    23: "qu-2025-zkgpt",
    24: "chantasantitam-2026-palm",
}

NUMBER_COL = 1  # A: Paper #
TITLE_COL = 2  # B: primary paper
THEME_COL = 3  # C: theme (vertically merged across its papers)
TOPIC_COL = 4  # D: topic (Part 1) or Software/Hardware (Part 2)
ESSENTIAL_COL = 5  # E: essential readings

UC_PAGE = "under-construction.md"
UC_MARK = ' <sup title="Reading companion under construction">&dagger;</sup>'

HEADERS = ("#", "Topic", "Reading")


@dataclass(frozen=True)
class XlsxCell:
    text: str
    href: str | None


@dataclass(frozen=True)
class Reading:
    title: str
    href: str | None


@dataclass
class Paper:
    number: int
    title: str
    topic: str
    theme: str
    essential: list[Reading] = field(default_factory=list)


@dataclass
class Part:
    title: str
    papers: list[Paper] = field(default_factory=list)


def read_sheet(
    path: Path, sheet: str
) -> tuple[dict[tuple[int, int], XlsxCell], int, int]:
    """Return every non-empty cell as a ``{(row, col): XlsxCell}`` map.

    All openpyxl access is confined here; values cross the boundary as plain
    strings via :func:`cast`, so the rest of the module stays statically typed.
    """
    workbook = load_workbook(path)
    worksheet = workbook[sheet]
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    cells: dict[tuple[int, int], XlsxCell] = {}
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            value = cast("object | None", cell.value)
            link = cell.hyperlink
            href = link.target if link is not None else None
            text = "" if value is None else " ".join(str(value).split())
            if text or href:
                cells[(row, col)] = XlsxCell(text=text, href=href)
    return cells, max_row, max_col


def cell_text(cells: dict[tuple[int, int], XlsxCell], row: int, col: int) -> str:
    cell = cells.get((row, col))
    return cell.text if cell else ""


def readings(
    cells: dict[tuple[int, int], XlsxCell], rows: range, cols: range
) -> list[Reading]:
    out: list[Reading] = []
    for row in rows:
        for col in cols:
            cell = cells.get((row, col))
            if cell and cell.text:
                out.append(Reading(title=cell.text, href=cell.href))
    return out


def parse(cells: dict[tuple[int, int], XlsxCell], max_row: int) -> list[Part]:
    """Turn the cell map into Part -> Paper structure using cell values alone."""
    part_rows = [
        r
        for r in range(2, max_row + 1)
        if cell_text(cells, r, NUMBER_COL).startswith("Part")
    ]
    paper_rows = [
        r for r in range(2, max_row + 1) if cell_text(cells, r, NUMBER_COL).isdigit()
    ]
    boundaries = sorted(part_rows + paper_rows + [max_row + 1])

    parts: list[Part] = []
    current_part: Part | None = None
    current_theme = ""
    for row in range(2, max_row + 1):
        head = cell_text(cells, row, NUMBER_COL)
        theme = cell_text(cells, row, THEME_COL)
        if theme:
            current_theme = theme
        if head.startswith("Part"):
            current_part = Part(title=head)
            parts.append(current_part)
            continue
        if not head.isdigit():
            continue
        if current_part is None:
            msg = f"paper row {row} appears before any Part header"
            raise ValueError(msg)
        end = min(b for b in boundaries if b > row) - 1
        block = range(row, end + 1)
        current_part.papers.append(
            Paper(
                number=int(head),
                title=cell_text(cells, row, TITLE_COL),
                topic=cell_text(cells, row, TOPIC_COL),
                theme=current_theme,
                essential=readings(
                    cells, block, range(ESSENTIAL_COL, ESSENTIAL_COL + 1)
                ),
            )
        )
    return parts


def esc(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def link(title: str, href: str | None) -> str:
    safe = esc(title)
    if not href:
        return safe
    return f'<a href="{href.replace("&", "&amp;")}">{safe}</a>'


def essential_details(items: list[Reading]) -> list[str]:
    """The collapsible essential-readings block; empty when there are none."""
    if not items:
        return []
    lines = [
        "        <details>",
        f"          <summary>Essential readings ({len(items)})</summary>",
        "          <ul>",
    ]
    lines += [f"            <li>{link(r.title, r.href)}</li>" for r in items]
    lines += ["          </ul>", "        </details>"]
    return lines


def paper_cell(paper: Paper) -> str:
    slug = READY.get(paper.number)
    if slug:
        return f'<a href="papers/{slug}.md">{esc(paper.title)}</a>'
    return f'<a href="{UC_PAGE}">{esc(paper.title)}</a>{UC_MARK}'


def part_heading(title: str) -> str:
    """Render 'Part 1 Risks ...' as 'Part 1: Risks ...'."""
    words = title.split(" ", 2)
    if len(words) == 3 and words[0] == "Part":
        return f"Part {words[1]}: {words[2]}"
    return title


def render_table(part: Part) -> str:
    lines = ["<table>", "  <thead>", "    <tr>"]
    lines += [f'      <th scope="col">{h}</th>' for h in HEADERS]
    lines += ["    </tr>", "  </thead>", "  <tbody>"]

    prev_theme: str | None = None
    for paper in part.papers:
        if paper.theme != prev_theme:
            lines.append("    <tr>")
            lines.append(
                f'      <th colspan="3" scope="colgroup">{esc(paper.theme)}</th>'
            )
            lines.append("    </tr>")
            prev_theme = paper.theme
        lines.append("    <tr>")
        lines.append(f"      <td>{paper.number}</td>")
        lines.append(f"      <td>{esc(paper.topic)}</td>")
        lines.append("      <td>")
        lines.append("        <strong>Assigned reading</strong>")
        lines.append(f"        {paper_cell(paper)}")
        lines += essential_details(paper.essential)
        lines.append("      </td>")
        lines.append("    </tr>")

    lines += ["  </tbody>", "</table>"]
    return "\n".join(lines)


def main() -> None:
    cells, max_row, _ = read_sheet(XLSX, SHEET)
    parts = parse(cells, max_row)
    blocks: list[str] = []
    for part in parts:
        blocks.append(f"### {part_heading(part.title)}")
        blocks.append(render_table(part))
    blocks.append(
        "<sup>&dagger;</sup> Reading companion under construction; the link opens"
        " a placeholder page."
    )
    print("\n\n".join(blocks))


if __name__ == "__main__":
    main()
